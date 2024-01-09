import os
import pickle
import csv
import openpyxl


def get_current_directory():
    return os.getcwd()

def load_summonerName_puuid_map(file_path = f'{get_current_directory()}/summonerName_puuid_map.pkl'):
    try:
        with open(file_path, 'rb') as file:
            summonerName_puuid_map = pickle.load(file)
        # print(f'LOAD: summonerName_puuid_map: {summonerName_puuid_map.keys()}')
        # print(f"summonerName_puuid map loaded from file: {file_path}")
        return summonerName_puuid_map
    except FileNotFoundError:
        # print(f"File not found: {file_path}")
        return {}

def save_summonerName_puuid_map(summonerName_puuid_map, file_path = f'{get_current_directory()}/summonerName_puuid_map.pkl'):
    with open(file_path, 'wb') as file:
        pickle.dump(summonerName_puuid_map, file)
    # print(f'SAVE: summonerName_puuid_map: {summonerName_puuid_map.keys()}')
    # print(f"summonerName_puuid map saved to file: {file_path}")

def save_match_data_map(match_data_map, file_path):
    # print(f'save: {match_data_map.keys()}')
    with open(file_path, 'wb') as file:
        pickle.dump(match_data_map, file)
    # print(f"Match data map saved to file: {file_path}")

def load_match_data_map(file_path):
    try:
        with open(file_path, 'rb') as file:
            match_data_map = pickle.load(file)
        # print(f'load: {match_data_map.keys()}')
        # print(f"Match data map loaded from file: {file_path}")
        return match_data_map
    except FileNotFoundError:
        # print(f"File not found: {file_path}")
        return {}

def create_excel_file(data, filename):
    """
    Creates an Excel file with the provided data.

    Args:
        data (list): List of dictionaries where each dictionary represents a row in the Excel file.
        filename (str): Name of the Excel file to be created.

    Returns:
        None
    """
    filename = filename + '.xlsx'
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    # Extract the keys from the first dictionary to use as column headers
    fieldnames = list(data[0].keys())
    
    # Write the column headers to the worksheet
    for col_num, fieldname in enumerate(fieldnames, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = fieldname
    
    # Write each dictionary as a row in the worksheet
    for row_num, row_data in enumerate(data, 2):
        for col_num, fieldname in enumerate(fieldnames, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = row_data[fieldname]
    
    # Save the workbook as an Excel file
    workbook.save(filename)
    
    print(f"Excel file '{filename}' created successfully.")

def create_csv_file(data, filename):
    """
    Creates a CSV file with the provided data.

    Args:
        data (list): List of dictionaries where each dictionary represents a row in the CSV file.
        filename (str): Name of the CSV file to be created.

    Returns:
        None
    """
    filename = filename + '.csv'
    # Extract the keys from the first dictionary to use as column headers
    fieldnames = data[0].keys()
    
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        # Write the column headers to the CSV file
        writer.writeheader()
        
        # Write each dictionary as a row in the CSV file
        writer.writerows(data)
    
    print(f"CSV file '{filename}' created successfully.")

def save_dataframe(df, csv_path, excel_path):
    # Save DataFrame as a CSV file
    df.to_csv(csv_path, index=False)
    
    # Save DataFrame as an Excel file
    df.to_excel(excel_path, index=False)
    
    print("DataFrame saved as CSV and Excel files.")

def remove_empty_dicts(data):
    """
    Removes empty dictionaries from a list of dictionaries.

    Args:
        data (list): List of dictionaries.

    Returns:
        list: List of dictionaries without empty dictionaries.
    """
    return [item for item in data if item]

# def load_summoner_puuid_matchHistory_map(file_path = f'{get_current_directory()}/summoner_puuid_matchHistory_map.pkl'):
#     try:
#         with open(file_path, 'rb') as file:
#             summoner_puuid_matchHistory_map = pickle.load(file)
#         # print(f'LOAD: summoner_puuid_matchHistory_map: {summoner_puuid_matchHistory_map.keys()}')
#         # print(f"summoner_puuid_matchHistory map loaded from file: {file_path}")
#         return summoner_puuid_matchHistory_map
#     except FileNotFoundError:
#         # print(f"File not found: {file_path}")
#         return {}

# def save_summoner_puuid_matchHistory_map(summoner_puuid_matchHistory_map, file_path = f'{get_current_directory()}/summoner_puuid_matchHistory_map.pkl'):
#     with open(file_path, 'wb') as file:
#         pickle.dump(summoner_puuid_matchHistory_map, file)
#     # print(f'SAVE: summoner_puuid_matchHistory_map: {summoner_puuid_matchHistory_map.keys()}')
#     # print(f"summoner_puuid_matchHistory map saved to file: {file_path}")